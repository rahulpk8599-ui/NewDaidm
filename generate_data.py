"""
LifePass Dubai — Perfect Synthetic Dataset Generator
1,000 respondents · Zero missing values · Realistic Dubai spend ranges
4 latent segments · Noise 5% · Outliers 3% · Right-skewed distributions
"""

import numpy as np
import pandas as pd
import warnings
warnings.filterwarnings("ignore")

SEED = 42
rng  = np.random.default_rng(SEED)
np.random.seed(SEED)
N    = 1000

SEGMENTS = {
    "Busy Professional": 280,
    "Nuclear Family":    300,
    "Budget Expat":      250,
    "Premium Resident":  170,
}
seg_labels = []
for name, count in SEGMENTS.items():
    seg_labels.extend([name] * count)
rng.shuffle(seg_labels)
seg_labels = np.array(seg_labels)

def clamp(v, lo, hi):   return float(np.clip(v, lo, hi))
def iclamp(v, lo, hi):  return int(np.clip(round(v), lo, hi))
def lognorm_sample(mean, sd_frac=0.30):
    sigma = np.sqrt(np.log(1 + sd_frac**2))
    mu    = np.log(mean) - sigma**2 / 2
    return float(rng.lognormal(mu, sigma))
def likert(mean, sd):   return iclamp(rng.normal(mean, sd), 1, 5)
def pick(choices, probs): return rng.choice(choices, p=probs)

# Spend params: (mean, sd_frac, min, max) - service spend only, not rent/school
SPEND_PARAMS = {
    "Budget Expat":      (480,   0.35,  200,   900),
    "Busy Professional": (1100,  0.40,  450,  2800),
    "Nuclear Family":    (2200,  0.45,  700,  5500),
    "Premium Resident":  (5500,  0.50, 1800, 15000),
}
INC_PROBS = {
    "Budget Expat":      [0.30, 0.42, 0.20, 0.06, 0.02],
    "Busy Professional": [0.02, 0.15, 0.45, 0.28, 0.10],
    "Nuclear Family":    [0.05, 0.22, 0.42, 0.22, 0.09],
    "Premium Resident":  [0.00, 0.02, 0.12, 0.36, 0.50],
}
INC_RANGES = {1:(2000,4800),2:(5000,9800),3:(10000,19500),4:(20000,34500),5:(35000,85000)}
WTP_PARAMS = {
    "Budget Expat":      (180,  0.35,  50,  350),
    "Busy Professional": (280,  0.35, 100,  550),
    "Nuclear Family":    (340,  0.35, 150,  650),
    "Premium Resident":  (580,  0.40, 200, 1200),
}
AGE_P = {"Busy Professional":(31,4),"Nuclear Family":(38,5),"Budget Expat":(28,4),"Premium Resident":(43,6)}
APPS_P = {"Busy Professional":(7.2,1.4),"Nuclear Family":(6.8,1.5),"Budget Expat":(4.6,1.4),"Premium Resident":(6.4,1.7)}
HOURS_P = {"Busy Professional":(3.9,0.8),"Nuclear Family":(3.5,0.9),"Budget Expat":(2.3,0.9),"Premium Resident":(2.7,1.0)}
SAT_P = {"Busy Professional":(2.1,0.7),"Nuclear Family":(2.4,0.8),"Budget Expat":(2.9,0.9),"Premium Resident":(2.3,0.8)}
PAIN_P = {"Busy Professional":(7.9,1.1),"Nuclear Family":(7.2,1.2),"Budget Expat":(6.0,1.5),"Premium Resident":(6.8,1.3)}
PROB_FREQ_P = {"Busy Professional":(3.6,0.8),"Nuclear Family":(3.3,0.9),"Budget Expat":(2.8,0.9),"Premium Resident":(3.0,1.0)}
IMP_ONE_APP_P = {"Busy Professional":(4.4,0.6),"Nuclear Family":(4.1,0.7),"Budget Expat":(3.4,0.9),"Premium Resident":(3.9,0.8)}
AI_P = {"Busy Professional":(4.1,0.7),"Nuclear Family":(3.3,0.9),"Budget Expat":(2.8,0.9),"Premium Resident":(4.0,0.8)}
SUBS_EXIST_P = {
    "Busy Professional":[0.04,0.18,0.42,0.36],
    "Nuclear Family":[0.09,0.30,0.38,0.23],
    "Budget Expat":[0.32,0.40,0.20,0.08],
    "Premium Resident":[0.02,0.10,0.33,0.55],
}
LIFESCORE_P = {"Busy Professional":(2.4,0.8),"Nuclear Family":(2.7,0.8),"Budget Expat":(3.1,0.9),"Premium Resident":(3.1,0.9)}
INTENT_P = {"Busy Professional":(4.2,0.7),"Nuclear Family":(4.0,0.8),"Budget Expat":(3.1,1.0),"Premium Resident":(4.1,0.8)}

NATS = ["Indian","Pakistani","Filipino","Arab","British/Western","Emirati","Other"]
NAT_P = {
    "Busy Professional":[0.38,0.15,0.12,0.12,0.13,0.04,0.06],
    "Nuclear Family":[0.42,0.20,0.14,0.10,0.06,0.04,0.04],
    "Budget Expat":[0.35,0.25,0.18,0.08,0.04,0.02,0.08],
    "Premium Resident":[0.28,0.10,0.06,0.16,0.22,0.10,0.08],
}
GENDER_OPTS = ["Male","Female","Prefer not to say"]
GENDER_PROBS = {
    "Budget Expat":[0.56,0.41,0.03],"Busy Professional":[0.52,0.45,0.03],
    "Nuclear Family":[0.51,0.46,0.03],"Premium Resident":[0.50,0.47,0.03],
}
LIVING = ["Living alone","Couple no kids","Family with kids","Flatmates","Extended family"]
LIVING_P = {
    "Busy Professional":[0.44,0.28,0.10,0.14,0.04],
    "Nuclear Family":[0.02,0.16,0.74,0.04,0.04],
    "Budget Expat":[0.18,0.14,0.20,0.40,0.08],
    "Premium Resident":[0.16,0.30,0.46,0.04,0.04],
}
AREAS = ["JLT/Marina","Downtown/Business Bay","Deira/Bur Dubai",
         "Mirdif/Sports City","International City/Al Quoz","Palm/Emirates Hills","Other"]
AREA_P = {
    "Busy Professional":[0.36,0.28,0.10,0.12,0.05,0.04,0.05],
    "Nuclear Family":[0.10,0.14,0.16,0.40,0.12,0.04,0.04],
    "Budget Expat":[0.05,0.05,0.28,0.14,0.40,0.02,0.06],
    "Premium Resident":[0.14,0.22,0.04,0.12,0.02,0.38,0.08],
}
FRUSTRATIONS = ["Too expensive","Too many apps to manage","Unreliable quality/timing",
                "No personalisation","Hard to book and track","No cross-service loyalty"]
FRUST_P = {
    "Busy Professional":[0.12,0.40,0.22,0.14,0.08,0.04],
    "Nuclear Family":[0.20,0.28,0.24,0.10,0.12,0.06],
    "Budget Expat":[0.46,0.18,0.18,0.06,0.08,0.04],
    "Premium Resident":[0.05,0.16,0.14,0.32,0.08,0.25],
}
PAY_MODES = ["Monthly subscription","Annual subscription","Pay per use + discount","Flexible credits"]
PAY_P = {
    "Busy Professional":[0.45,0.30,0.16,0.09],
    "Nuclear Family":[0.42,0.22,0.24,0.12],
    "Budget Expat":[0.26,0.08,0.48,0.18],
    "Premium Resident":[0.28,0.46,0.16,0.10],
}
TRUST_DRIVERS = ["Transparent pricing","Verified partners","Free trial",
                 "Peer recommendation","Cancellation policy","Data privacy"]
TRUST_P = {
    "Busy Professional":[0.22,0.20,0.24,0.16,0.10,0.08],
    "Nuclear Family":[0.20,0.24,0.20,0.22,0.08,0.06],
    "Budget Expat":[0.36,0.16,0.26,0.12,0.06,0.04],
    "Premium Resident":[0.12,0.32,0.12,0.16,0.08,0.20],
}
SVC_USE_P = {
    "Busy Professional":[0.84,0.80,0.76,0.82,0.72,0.44,0.26,0.56,0.34],
    "Nuclear Family":[0.92,0.94,0.64,0.90,0.70,0.74,0.38,0.32,0.22],
    "Budget Expat":[0.58,0.52,0.44,0.66,0.48,0.36,0.14,0.20,0.08],
    "Premium Resident":[0.76,0.88,0.82,0.74,0.82,0.80,0.46,0.85,0.72],
}
SVC_FREQ_MU = {
    "Busy Professional":[3.9,3.5,3.6,4.1,3.2,2.1,1.7],
    "Nuclear Family":[4.3,4.1,3.1,4.4,3.0,3.3,1.9],
    "Budget Expat":[2.7,2.4,2.1,3.3,2.3,1.8,1.3],
    "Premium Resident":[3.5,3.9,4.0,3.6,3.9,3.6,2.3],
}
CAR_P = {
    "Budget Expat":(["Own car","Family car","No car"],[0.18,0.12,0.70]),
    "Busy Professional":(["Own car","Family car","No car"],[0.56,0.20,0.24]),
    "Nuclear Family":(["Own car","Family car","No car"],[0.46,0.32,0.22]),
    "Premium Resident":(["Own car","Family car","No car"],[0.68,0.24,0.08]),
}
PET_P = {
    "Budget Expat":(["No pets","One pet","Multiple pets"],[0.74,0.20,0.06]),
    "Busy Professional":(["No pets","One pet","Multiple pets"],[0.64,0.28,0.08]),
    "Nuclear Family":(["No pets","One pet","Multiple pets"],[0.58,0.32,0.10]),
    "Premium Resident":(["No pets","One pet","Multiple pets"],[0.48,0.34,0.18]),
}

rows = []
for i, seg in enumerate(seg_labels):
    age    = iclamp(rng.normal(*AGE_P[seg]), 18, 65)
    gender = pick(GENDER_OPTS, GENDER_PROBS[seg])
    nationality = pick(NATS, NAT_P[seg])
    inc_ord = int(pick([1,2,3,4,5], INC_PROBS[seg]))
    lo, hi  = INC_RANGES[inc_ord]
    income  = round(float(rng.uniform(lo, hi)), -2)
    living  = pick(LIVING, LIVING_P[seg])
    area    = pick(AREAS, AREA_P[seg])
    n_apps  = iclamp(rng.normal(*APPS_P[seg]), 1, 14)

    mu_sp, sd_frac, sp_min, sp_max = SPEND_PARAMS[seg]
    inc_factor = float(np.clip(1 + 0.15*(income-15000)/15000, 0.5, 2.5))
    spend = round(float(np.clip(lognorm_sample(mu_sp*inc_factor, sd_frac), sp_min, sp_max)), -1)

    hours_ord    = likert(*HOURS_P[seg])
    svc_used     = [int(rng.random() < p) for p in SVC_USE_P[seg]]
    satisfaction = likert(*SAT_P[seg])
    prob_freq    = likert(*PROB_FREQ_P[seg])

    pain_base  = rng.normal(*PAIN_P[seg])
    pain_score = round(float(np.clip(
        pain_base + 0.14*(n_apps-6) - 0.28*(satisfaction-2.5), 1.0, 10.0)), 1)

    frustration = pick(FRUSTRATIONS, FRUST_P[seg])
    imp_one_app = likert(*IMP_ONE_APP_P[seg])
    svc_freqs   = [likert(fmu, 0.85) for fmu in SVC_FREQ_MU[seg]]

    opts, probs = CAR_P[seg]; car = pick(opts, probs)
    opts, probs = PET_P[seg]; pet = pick(opts, probs)

    subs_exist  = int(pick([0,1,2,3], SUBS_EXIST_P[seg]))
    intent_raw  = rng.normal(*INTENT_P[seg])
    intent      = iclamp(intent_raw + 0.10*(pain_score-6) + 0.10*(imp_one_app-3) + 0.08*(subs_exist-1), 1, 5)

    mu_w, sd_frac_w, wtp_min, wtp_max = WTP_PARAMS[seg]
    wtp = round(float(np.clip(
        lognorm_sample(mu_w, sd_frac_w) + 0.003*(income-15000) + 6.0*(pain_score-6.0),
        wtp_min, wtp_max)), -1)

    if   wtp < 100:  wtp_cat = "Below AED 100"
    elif wtp < 200:  wtp_cat = "AED 100-199"
    elif wtp < 300:  wtp_cat = "AED 200-299"
    elif wtp < 400:  wtp_cat = "AED 300-399"
    elif wtp < 500:  wtp_cat = "AED 400-499"
    elif wtp < 750:  wtp_cat = "AED 500-749"
    else:            wtp_cat = "AED 750+"

    pay_mode     = pick(PAY_MODES, PAY_P[seg])
    svc_weights  = np.array(svc_freqs[:7]+[svc_used[7]*3, svc_used[8]*2], dtype=float)
    svc_weights += rng.uniform(0, 0.4, len(svc_weights))
    top3_idx     = np.argsort(svc_weights)[::-1][:3]
    svc_ranks    = [0]*9
    for rp, si in enumerate(top3_idx): svc_ranks[si] = rp+1

    ai_comfort   = likert(*AI_P[seg])
    trust_driver = pick(TRUST_DRIVERS, TRUST_P[seg])
    lifescore    = likert(*LIFESCORE_P[seg])

    sub_score    = ((intent/5)*0.40 + (min(wtp,600)/600)*0.25 +
                   ((5-satisfaction)/4)*0.20 + (ai_comfort/5)*0.15)
    would_sub    = int(sub_score >= 0.48)

    rows.append({
        "Respondent_ID":f"LP{i+1:04d}","Segment":seg,"Gender":gender,
        "Q1_Age":age,"Q2_Nationality":nationality,
        "Q3_Income_Bracket":inc_ord,"Q3_Income_AED_Month":income,
        "Q4_Living_Situation":living,"Q5_Area_Dubai":area,
        "Q6_Num_Apps_Used":n_apps,"Q7_Monthly_Service_Spend_AED":spend,
        "Q8_Hours_Logistics_Per_Week":hours_ord,
        "Q9_Uses_Laundry":svc_used[0],"Q9_Uses_Cleaning":svc_used[1],
        "Q9_Uses_Carwash":svc_used[2],"Q9_Uses_Grocery":svc_used[3],
        "Q9_Uses_Salon":svc_used[4],"Q9_Uses_Maintenance":svc_used[5],
        "Q9_Uses_Petcare":svc_used[6],"Q9_Uses_Airport":svc_used[7],
        "Q9_Uses_Concierge":svc_used[8],
        "Q10_Satisfaction_Current":satisfaction,
        "Q11_Problem_Frequency":prob_freq,"Q12_Pain_Score_1to10":pain_score,
        "Q13_Primary_Frustration":frustration,"Q14_Importance_One_App":imp_one_app,
        "Q15_Freq_Laundry":svc_freqs[0],"Q15_Freq_Cleaning":svc_freqs[1],
        "Q15_Freq_Carwash":svc_freqs[2],"Q15_Freq_Grocery":svc_freqs[3],
        "Q15_Freq_Salon":svc_freqs[4],"Q15_Freq_Maintenance":svc_freqs[5],
        "Q15_Freq_Petcare":svc_freqs[6],
        "Q16_Car_Ownership":car,"Q17_Pet_Ownership":pet,
        "Q18_Subscription_Intent":intent,"Q19_WTP_AED_Month":wtp,
        "Q19_WTP_Category":wtp_cat,"Q20_Preferred_Payment":pay_mode,
        "Q21_Rank_Laundry":svc_ranks[0],"Q21_Rank_Cleaning":svc_ranks[1],
        "Q21_Rank_Carwash":svc_ranks[2],"Q21_Rank_Grocery":svc_ranks[3],
        "Q21_Rank_Salon":svc_ranks[4],"Q21_Rank_Maintenance":svc_ranks[5],
        "Q21_Rank_Petcare":svc_ranks[6],"Q21_Rank_Airport":svc_ranks[7],
        "Q21_Rank_Concierge":svc_ranks[8],
        "Q22_AI_Comfort":ai_comfort,"Q23_Existing_Subscriptions":subs_exist,
        "Q24_Trust_Driver":trust_driver,"Q25_LifeScore":lifescore,
        "Would_Subscribe":would_sub,
    })

df = pd.DataFrame(rows)

# ── Noise (5%)
NOISE_COLS = ["Q10_Satisfaction_Current","Q14_Importance_One_App",
              "Q18_Subscription_Intent","Q22_AI_Comfort","Q25_LifeScore"]
for idx in rng.choice(N, size=int(N*0.05), replace=False):
    df.at[idx, rng.choice(NOISE_COLS)] = int(rng.integers(1,6))
    df.at[idx,"Q6_Num_Apps_Used"] = iclamp(df.at[idx,"Q6_Num_Apps_Used"]+rng.integers(-2,3),1,14)

# ── Outliers (3%)
for idx in rng.choice(N, size=int(N*0.03), replace=False):
    t = rng.integers(0,4)
    if   t==0: df.at[idx,"Q7_Monthly_Service_Spend_AED"] = round(float(rng.uniform(8000,15000)),-2)
    elif t==1: df.at[idx,"Q19_WTP_AED_Month"]=round(float(rng.uniform(900,1200)),-1); df.at[idx,"Q19_WTP_Category"]="AED 750+"
    elif t==2: df.at[idx,"Q3_Income_AED_Month"]=round(float(rng.uniform(60000,90000)),-3); df.at[idx,"Q3_Income_Bracket"]=5
    else:      df.at[idx,"Q12_Pain_Score_1to10"]=round(float(rng.uniform(9.3,10.0)),1); df.at[idx,"Q6_Num_Apps_Used"]=int(rng.integers(11,15))

# ── Recalculate target
sub_score = ((df["Q18_Subscription_Intent"]/5)*0.40 +
             (df["Q19_WTP_AED_Month"].clip(0,600)/600)*0.25 +
             ((5-df["Q10_Satisfaction_Current"])/4)*0.20 +
             (df["Q22_AI_Comfort"]/5)*0.15)
df["Would_Subscribe"] = (sub_score >= sub_score.quantile(0.32)).astype(int)

# ── Shuffle
df = df.sample(frac=1, random_state=SEED).reset_index(drop=True)
df["Respondent_ID"] = [f"LP{i+1:04d}" for i in range(N)]

assert df.isnull().sum().sum() == 0, "Missing values found!"

# ── Save CSV
csv_path = "/home/claude/lifepass_app/lifepass_survey_1000.csv"
df.to_csv(csv_path, index=False)

# ── Save Excel
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

xl_path = "/home/claude/lifepass_app/lifepass_survey_1000.xlsx"
with pd.ExcelWriter(xl_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Survey Data")
    ws = writer.sheets["Survey Data"]
    hfill = PatternFill("solid", fgColor="1B3A6B")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ts = Side(style="thin", color="CCCCCC")
    tb = Border(left=ts, right=ts, top=ts, bottom=ts)
    for ci, cn in enumerate(df.columns, 1):
        cell = ws.cell(1, ci)
        cell.fill=hfill; cell.font=hfont; cell.alignment=halign; cell.border=tb
    for ci in range(1, len(df.columns)+1):
        ml = max(len(df.columns[ci-1]), df.iloc[:,ci-1].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(ml+2, 30)
    ws.freeze_panes = "A2"
    sfill = PatternFill("solid", fgColor="EFF6FF")
    for ri in range(2, N+2, 2):
        for ci in range(1, len(df.columns)+1):
            ws.cell(ri,ci).fill = sfill
    # Summary sheet
    ws2 = writer.book.create_sheet("Summary Stats")
    sumdata = [
        ("Total Respondents", len(df)),
        ("Missing Values", df.isnull().sum().sum()),
        ("Would Subscribe (%)", f"{df['Would_Subscribe'].mean()*100:.1f}%"),
        ("Would NOT Subscribe (%)", f"{(1-df['Would_Subscribe'].mean())*100:.1f}%"),
        ("Spend — Min (AED)", f"{df['Q7_Monthly_Service_Spend_AED'].min():.0f}"),
        ("Spend — Median (AED)", f"{df['Q7_Monthly_Service_Spend_AED'].median():.0f}"),
        ("Spend — Mean (AED)", f"{df['Q7_Monthly_Service_Spend_AED'].mean():.0f}"),
        ("Spend — P90 (AED)", f"{df['Q7_Monthly_Service_Spend_AED'].quantile(0.9):.0f}"),
        ("Spend — Max (AED)", f"{df['Q7_Monthly_Service_Spend_AED'].max():.0f}"),
        ("Spend Skewness", f"{df['Q7_Monthly_Service_Spend_AED'].skew():.3f}"),
        ("WTP — Median (AED)", f"{df['Q19_WTP_AED_Month'].median():.0f}"),
        ("WTP — Max (AED)", f"{df['Q19_WTP_AED_Month'].max():.0f}"),
        ("Income — Mean (AED)", f"{df['Q3_Income_AED_Month'].mean():.0f}"),
        ("Income — Max (AED)", f"{df['Q3_Income_AED_Month'].max():.0f}"),
        ("Income Skewness", f"{df['Q3_Income_AED_Month'].skew():.3f}"),
        ("Avg Pain Score", f"{df['Q12_Pain_Score_1to10'].mean():.2f}/10"),
    ]
    for ri, (k, v) in enumerate(sumdata, 1):
        ws2.cell(ri,1,k).font = Font(bold=True, color="1B3A6B")
        ws2.cell(ri,2,str(v))
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
    # Segment spend sheet
    ws3 = writer.book.create_sheet("Spend by Segment")
    seg_spend = df.groupby("Segment")["Q7_Monthly_Service_Spend_AED"].describe()
    seg_spend.to_excel(writer, sheet_name="Spend by Segment")

# ── Print summary
print("="*65)
print("  LifePass Dubai — Final Dataset")
print("="*65)
print(f"  Rows : {len(df)} | Cols : {len(df.columns)} | Missing : {df.isnull().sum().sum()} ✓")
print()
print("  Segments:")
for s,c in df["Segment"].value_counts().items(): print(f"    {s:<25}: {c} ({c/N*100:.1f}%)")
print()
vc=df["Would_Subscribe"].value_counts()
print(f"  Target: Subscribe {vc[1]} ({vc[1]/N*100:.1f}%) | Won't {vc[0]} ({vc[0]/N*100:.1f}%)")
print()
sp=df["Q7_Monthly_Service_Spend_AED"]
print("  Service Spend (AED):")
print(f"    Min={sp.min():.0f}  P25={sp.quantile(.25):.0f}  Median={sp.median():.0f}  "
      f"Mean={sp.mean():.0f}  P90={sp.quantile(.9):.0f}  Max={sp.max():.0f}  Skew={sp.skew():.2f}")
print()
print("  Spend by segment (mean | max):")
for s in ["Budget Expat","Busy Professional","Nuclear Family","Premium Resident"]:
    sub=df[df["Segment"]==s]["Q7_Monthly_Service_Spend_AED"]
    print(f"    {s:<25}: mean AED {sub.mean():>6,.0f}  |  max AED {sub.max():>7,.0f}")
print()
inc=df["Q3_Income_AED_Month"]
print(f"  Income: mean AED {inc.mean():,.0f} | max AED {inc.max():,.0f} | skew {inc.skew():.2f}")
wtp=df["Q19_WTP_AED_Month"]
print(f"  WTP   : median AED {wtp.median():.0f} | max AED {wtp.max():.0f}")
print()
print(f"  CSV  saved → {csv_path}")
print(f"  XLSX saved → {xl_path}")
print("="*65)
